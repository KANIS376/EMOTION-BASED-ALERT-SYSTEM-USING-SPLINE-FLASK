import cv2
import numpy as np
from deepface import DeepFace
from flask import Flask, render_template, Response, request, jsonify, session, redirect, url_for, send_file
import threading
import os
import json
import math
from passlib.hash import pbkdf2_sha256
from uuid import uuid4
from datetime import datetime
import logging
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.urandom(24)  # Consistent secret key for session persistence

# Setup logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Emotion labels
emotion_labels = ['angry', 'disgust', 'fear', 'happy', 'sad', 'surprise', 'neutral']

# Admin hashed password (rbm123)
ADMIN_PASSWORD_HASH = pbkdf2_sha256.hash("rbm123")

# Load Haar Cascade for face detection
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

# Global variables for video capture and threading
cap = None
frame = None
lock = threading.Lock()
stop_camera = False
latest_emotion = {'timestamp': datetime.now().isoformat(), 'emotion': 'none', 'confidence': 0.0}

# Paths
USERS_FILE = 'users.json'
EMOTIONS_DIR = 'data/emotions'

# Initialize users.json and emotions directory
if not os.path.exists(USERS_FILE):
    logger.debug(f"Creating users file: {USERS_FILE}")
    with open(USERS_FILE, 'w') as f:
        json.dump({}, f)

if not os.path.exists(EMOTIONS_DIR):
    logger.debug(f"Creating emotions directory: {EMOTIONS_DIR}")
    os.makedirs(EMOTIONS_DIR)

def load_users():
    try:
        with open(USERS_FILE, 'r') as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Error loading users from {USERS_FILE}: {e}")
        return {}

def save_users(users):
    try:
        with open(USERS_FILE, 'w') as f:
            json.dump(users, f, indent=4)
        logger.debug(f"Saved users to {USERS_FILE}")
    except Exception as e:
        logger.error(f"Error saving users to {USERS_FILE}: {e}")

def load_emotions(username):
    emotion_file = os.path.join(EMOTIONS_DIR, f"{username}_emotions.json")
    try:
        if os.path.exists(emotion_file):
            with open(emotion_file, 'r') as f:
                data = json.load(f)
                
                # Flag to track if we need to update the file
                updated = False
                
                # Ensure all sessions have all warning count fields
                for session in data:
                    # Check for sadness warning count
                    if 'warning_count' not in session:
                        session['warning_count'] = 0
                        updated = True
                    
                    # Check for fear warning count
                    if 'fear_warning_count' not in session:
                        session['fear_warning_count'] = 0
                        updated = True
                    
                    # Check for disgust warning count
                    if 'disgust_warning_count' not in session:
                        session['disgust_warning_count'] = 0
                        updated = True
                    
                    # Check for anger warning count
                    if 'angry_warning_count' not in session:
                        session['angry_warning_count'] = 0
                        updated = True
                
                # Save the updated data back to the file to ensure warning counts are persisted
                if updated:
                    with open(emotion_file, 'w') as f:
                        json.dump(data, f, indent=4)
                    logger.info(f"Updated warning counts in {username}_emotions.json")
                
                logger.debug(f"Loaded emotions for {username}: {len(data)} sessions")
                return data
        logger.debug(f"No emotion file for {username}, returning empty list")
        return []
    except Exception as e:
        logger.error(f"Error loading emotions for {username} from {emotion_file}: {e}")
        return []

def calculate_emotion_averages(emotions):
    """
    Calculate average confidence values for each emotion from a list of emotion records.
    
    Args:
        emotions (list): List of emotion data dictionaries, each containing 'emotion' and 'confidence' keys
        
    Returns:
        dict: Dictionary with emotion names as keys and average confidence values (0-100) as values
    """
    # Initialize stats for each emotion
    emotion_stats = {emotion: {'count': 0, 'total_confidence': 0.0} for emotion in emotion_labels}
    
    # Skip processing if emotions list is empty
    if not emotions:
        return {emotion: 0.0 for emotion in emotion_labels}
    
    # Process each emotion record
    valid_records = 0
    for emotion_data in emotions:
        try:
            # Validate the emotion data
            if not isinstance(emotion_data, dict):
                logger.warning(f"Invalid emotion data format (not a dict): {type(emotion_data)}")
                continue
                
            if 'emotion' not in emotion_data or 'confidence' not in emotion_data:
                logger.warning(f"Missing required keys in emotion data: {emotion_data}")
                continue
                
            emotion = emotion_data['emotion']
            
            # Try to convert confidence to float if it's not already
            try:
                confidence = float(emotion_data['confidence'])
            except (ValueError, TypeError):
                logger.warning(f"Invalid confidence value: {emotion_data['confidence']}")
                continue
                
            # Skip invalid confidence values
            if math.isnan(confidence) or confidence < 0 or confidence > 100:
                logger.warning(f"Confidence value out of range (0-100): {confidence}")
                continue
                
            # Only process recognized emotions
            if emotion in emotion_stats:
                emotion_stats[emotion]['count'] += 1
                emotion_stats[emotion]['total_confidence'] += confidence
                valid_records += 1
            else:
                logger.warning(f"Unknown emotion: {emotion}")
        except Exception as e:
            logger.error(f"Error processing emotion data: {e}")
            continue
    
    # Calculate averages for each emotion
    emotion_averages = {}
    for emotion in emotion_labels:
        count = emotion_stats[emotion]['count']
        total_confidence = emotion_stats[emotion]['total_confidence']
        
        # Calculate average, ensuring we return percentages (0-100)
        try:
            if count > 0:
                avg = float(total_confidence / count)
                # Clamp value to valid range
                avg = max(0.0, min(100.0, avg))
            else:
                avg = 0.0
        except Exception as e:
            logger.error(f"Error calculating average for {emotion}: {e}")
            avg = 0.0
            
        emotion_averages[emotion] = avg
    
    logger.debug(f"Calculated emotion averages from {valid_records} valid records: {emotion_averages}")
    return emotion_averages

def create_user_excel_report(username, output_path=None):
    """
    Generate a detailed Excel report for a specific user with all their session data.
    
    Args:
        username (str): The username to generate the report for
        output_path (str, optional): Path to save the Excel file. If None, returns BytesIO object.
        
    Returns:
        BytesIO or None: If output_path is None, returns BytesIO object with Excel data
    """
    logger.info(f"Generating Excel report for user: {username}")
    
    # Load all emotion data for the user
    emotions = load_emotions(username)
    
    if not emotions:
        logger.warning(f"No emotion data found for user: {username}")
        return None
    
    # Sort sessions by start_time
    sorted_emotions = sorted(emotions, key=lambda x: x.get('start_time', '0'), reverse=False)
    
    # Create a dictionary to store session counts for each day
    session_counts = {}
    
    # Prepare data for the sessions sheet
    sessions_data = []
    for idx, sess in enumerate(sorted_emotions, 1):
        # Get the date from the start_time
        start_time = sess.get('start_time', 'N/A')
        if start_time != 'N/A':
            try:
                # Extract the date part (YYYY-MM-DD)
                session_date = start_time.split('T')[0]
                
                # Increment the session count for this date
                if session_date not in session_counts:
                    session_counts[session_date] = 1
                else:
                    session_counts[session_date] += 1
                
                # Format the session label as "Username - YYYY-MM-DD - Session #"
                session_label = f"{username} - {session_date} - Session {session_counts[session_date]}"
            except Exception as e:
                logger.error(f"Error parsing start_time {start_time}: {e}")
                session_label = f"{username} - Session {idx} (time unknown)"
        else:
            session_label = f"{username} - Session {idx} (time unknown)"
        
        # Calculate session duration
        duration = sess.get('duration_formatted', 'N/A')
        if duration == 'N/A' and 'start_time' in sess and 'end_time' in sess and sess['end_time']:
            try:
                start_dt = datetime.fromisoformat(sess['start_time'])
                end_dt = datetime.fromisoformat(sess['end_time'])
                duration_seconds = (end_dt - start_dt).total_seconds()
                minutes = int(duration_seconds // 60)
                seconds = int(duration_seconds % 60)
                duration = f"{minutes}m {seconds}s"
            except Exception as e:
                logger.error(f"Error calculating duration: {e}")
        
        # Create row for sessions sheet
        row = {
            'Session Number': idx,
            'Session ID': sess.get('session_id', 'Unknown'),
            'Session Label': session_label,
            'Start Time': start_time,
            'End Time': sess.get('end_time', 'N/A'),
            'Duration': duration,
            'Warning Count': sess.get('warning_count', 0),
            'Emotions Recorded': len(sess.get('emotions', [])),
            'Login Type': sess.get('login_type', 'regular'),
            'First Session': 'Yes' if sess.get('first_session', False) else 'No',
            'User Agent': sess.get('user_agent', 'Unknown'),
            'IP Address': sess.get('ip_address', 'Unknown')
        }
        
        # Add emotion averages
        averages = sess.get('emotion_averages', {})
        for emotion in emotion_labels:
            row[f'{emotion.capitalize()} Average'] = averages.get(emotion, 0.0)
        
        sessions_data.append(row)
    
    # Prepare data for the emotions sheet (detailed emotion records)
    emotions_data = []
    for idx, sess in enumerate(sorted_emotions, 1):
        session_id = sess.get('session_id', 'Unknown')
        session_label = f"Session {idx}"
        
        # Add each emotion record
        for emotion_record in sess.get('emotions', []):
            row = {
                'Session Number': idx,
                'Session ID': session_id,
                'Session Label': session_label,
                'Timestamp': emotion_record.get('timestamp', 'Unknown'),
                'Emotion': emotion_record.get('emotion', 'Unknown'),
                'Confidence': emotion_record.get('confidence', 0.0)
            }
            emotions_data.append(row)
    
    # Create DataFrames
    sessions_df = pd.DataFrame(sessions_data)
    emotions_df = pd.DataFrame(emotions_data)
    
    # Create Excel file
    if output_path:
        output = output_path
    else:
        output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write the sessions sheet
        sessions_df.to_excel(writer, index=False, sheet_name='Sessions', float_format='%.2f')
        
        # Write the emotions sheet
        emotions_df.to_excel(writer, index=False, sheet_name='Emotion Records', float_format='%.2f')
        
        # Get the workbook and the worksheets
        workbook = writer.book
        sessions_worksheet = writer.sheets['Sessions']
        emotions_worksheet = writer.sheets['Emotion Records']
        
        # Format the Sessions worksheet
        for idx, col in enumerate(sessions_df.columns, 1):
            column_letter = get_column_letter(idx)
            sessions_worksheet.column_dimensions[column_letter].width = 15
            
            # Format header
            cell = sessions_worksheet[f"{column_letter}1"]
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Format warning count column with conditional formatting
            if col == 'Warning Count':
                for row_idx, value in enumerate(sessions_df['Warning Count'], 2):
                    cell = sessions_worksheet[f"{column_letter}{row_idx}"]
                    if value > 5:
                        cell.fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
                    elif value > 0:
                        cell.fill = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
        
        # Format the Emotions worksheet
        for idx, col in enumerate(emotions_df.columns, 1):
            column_letter = get_column_letter(idx)
            emotions_worksheet.column_dimensions[column_letter].width = 15
            
            # Format header
            cell = emotions_worksheet[f"{column_letter}1"]
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D0E0E3", end_color="D0E0E3", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    logger.info(f"Excel report generated for {username} with {len(sessions_data)} sessions and {len(emotions_data)} emotion records")
    
    if output_path:
        return None
    else:
        output.seek(0)
        return output

def save_emotions(username, session_data):
    emotion_file = os.path.join(EMOTIONS_DIR, f"{username}_emotions.json")
    try:
        emotions = load_emotions(username)
        
        # Calculate emotion averages if emotions data exists
        if session_data.get('emotions', []):
            session_data['emotion_averages'] = calculate_emotion_averages(session_data['emotions'])
        else:
            session_data['emotion_averages'] = {emotion: 0.0 for emotion in emotion_labels}
            
        # Ensure required session fields exist
        if 'warning_count' not in session_data:
            session_data['warning_count'] = 0
            
        # Add timestamp for last update
        session_data['last_updated'] = datetime.now().isoformat()
        
        # Calculate session duration if both start and end times are available
        if 'start_time' in session_data and 'end_time' in session_data and session_data['end_time']:
            try:
                start_dt = datetime.fromisoformat(session_data['start_time'])
                end_dt = datetime.fromisoformat(session_data['end_time'])
                duration_seconds = (end_dt - start_dt).total_seconds()
                session_data['duration_seconds'] = duration_seconds
                
                # Format duration as minutes and seconds for display
                minutes = int(duration_seconds // 60)
                seconds = int(duration_seconds % 60)
                session_data['duration_formatted'] = f"{minutes}m {seconds}s"
            except Exception as e:
                logger.error(f"Error calculating session duration: {e}")
                session_data['duration_seconds'] = 0
                session_data['duration_formatted'] = 'N/A'
            
        existing_session = next((s for s in emotions if s['session_id'] == session_data['session_id']), None)
        if existing_session:
            # Preserve warning_count if it exists in the existing session but not in the new data
            if 'warning_count' in existing_session and 'warning_count' not in session_data:
                session_data['warning_count'] = existing_session['warning_count']
            emotions[emotions.index(existing_session)] = session_data
        else:
            emotions.append(session_data)
            
        # Keep only the most recent 100 sessions
        if len(emotions) > 100:
            emotions = sorted(emotions, key=lambda x: x.get('start_time', '0'), reverse=True)[:100]
            
        # Save to JSON file with pretty formatting
        with open(emotion_file, 'w') as f:
            json.dump(emotions, f, indent=4)
            
        logger.debug(f"Saved session for {username} to {emotion_file}: {session_data['session_id']}, "
                    f"Emotions: {len(session_data.get('emotions', []))}, "
                    f"Averages: {session_data['emotion_averages']}, "
                    f"Warning Count: {session_data.get('warning_count', 0)}, "
                    f"Duration: {session_data.get('duration_formatted', 'N/A')}")
    except Exception as e:
        logger.error(f"Error saving emotions for {username} to {emotion_file}: {e}")
        raise  # Re-raise to catch in calling routes

def generate_frames():
    global cap, frame, stop_camera, latest_emotion
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        logger.error("Failed to open webcam")
        return
    stop_camera = False
    while not stop_camera:
        success, current_frame = cap.read()
        if not success:
            logger.warning("Failed to read frame")
            break
        gray_frame = cv2.cvtColor(current_frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray_frame, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))
        emotion_data = {'timestamp': datetime.now().isoformat(), 'emotion': 'none', 'confidence': 0.0}
        for (x, y, w, h) in faces:
            cv2.rectangle(current_frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
            face_img = current_frame[y:y + h, x:x + w]
            try:
                analysis = DeepFace.analyze(
                    img_path=face_img,
                    actions=['emotion'],
                    enforce_detection=False
                )
                emotion = analysis[0]['dominant_emotion']
                confidence = analysis[0]['emotion'][emotion]
                cv2.putText(current_frame, f"{emotion} ({confidence:.1f}%)",
                            (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)
                emotion_data = {
                    'timestamp': datetime.now().isoformat(),
                    'emotion': emotion,
                    'confidence': float(confidence)
                }
                logger.debug(f"Detected emotion: {emotion_data}")
            except Exception as e:
                logger.error(f"Error in emotion analysis: {e}")
        with lock:
            latest_emotion = emotion_data
            frame = current_frame.copy()
        ret, buffer = cv2.imencode('.jpg', current_frame)
        if not ret:
            logger.warning("Failed to encode frame")
            continue
        frame_bytes = buffer.tobytes()
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')
    if cap is not None:
        cap.release()
        logger.debug("Camera released")

@app.route('/')
def home():
    logger.debug("Accessing home route")
    return render_template('index.html')

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session or session.get('is_admin', False):
        logger.warning("Unauthorized access to dashboard")
        return redirect(url_for('home'))
    logger.debug(f"User {session.get('username')} accessed dashboard")
    return render_template('dashboard.html', username=session.get('username'))

@app.route('/admin_dashboard')
def admin_dashboard():
    if not session.get('is_admin', False):
        logger.warning("Unauthorized access to admin_dashboard")
        return redirect(url_for('home'))
    users = load_users()
    user_data = [
        {'username': data['username'], 'email': email}
        for email, data in users.items()
    ]
    logger.debug(f"Rendering admin dashboard with {len(user_data)} users")
    return render_template('admin.html', users=user_data)

@app.route('/download_user_data/<username>')
def download_user_data(username):
    if not session.get('is_admin', False):
        logger.warning(f"Unauthorized attempt to download data for {username}")
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    emotion_file = os.path.join(EMOTIONS_DIR, f"{username}_emotions.json")
    if not os.path.exists(emotion_file):
        logger.error(f"Emotion file for {username} not found")
        return jsonify({'success': False, 'message': 'User data not found'}), 404
    try:
        # Load emotions data, which will also ensure warning_count is added if missing
        emotions = load_emotions(username)
        
        # Save any updates back to the file
        updated = False
        for sess in emotions:
            if 'warning_count' not in sess:
                sess['warning_count'] = 0
                updated = True
        
        if updated:
            with open(emotion_file, 'w') as f:
                json.dump(emotions, f, indent=4)
            logger.info(f"Updated warning_count in {username}_emotions.json before Excel export")
        
        data = []
        
        # Sort sessions by start_time to ensure consistent session numbering
        sorted_emotions = sorted(emotions, key=lambda x: x.get('start_time', '0'), reverse=False)
        
        # Create a dictionary to store session counts for each day
        session_counts = {}
        
        for sess in sorted_emotions:
            # Ensure warning_count exists in the session
            if 'warning_count' not in sess:
                sess['warning_count'] = 0
                updated = True
            
            # Get the date from the start_time
            start_time = sess.get('start_time', 'N/A')
            if start_time != 'N/A':
                try:
                    # Extract the date part (YYYY-MM-DD)
                    session_date = start_time.split('T')[0]
                    
                    # Increment the session count for this date
                    if session_date not in session_counts:
                        session_counts[session_date] = 1
                    else:
                        session_counts[session_date] += 1
                    
                    # Format the session label as "Username - YYYY-MM-DD - Session #"
                    session_label = f"{username} - {session_date} - Session {session_counts[session_date]}"
                except Exception as e:
                    logger.error(f"Error parsing start_time {start_time}: {e}")
                    session_label = f"{username} - Session (time unknown)"
            else:
                session_label = f"{username} - Session (time unknown)"
                
            # Calculate session duration if both start and end times are available
            start_time = sess.get('start_time', 'N/A')
            end_time = sess.get('end_time', 'N/A')
            duration = 'N/A'
            
            if start_time != 'N/A' and end_time != 'N/A':
                try:
                    start_dt = datetime.fromisoformat(start_time)
                    end_dt = datetime.fromisoformat(end_time)
                    duration_seconds = (end_dt - start_dt).total_seconds()
                    
                    # Format duration as minutes and seconds
                    minutes = int(duration_seconds // 60)
                    seconds = int(duration_seconds % 60)
                    duration = f"{minutes}m {seconds}s"
                except Exception as e:
                    logger.error(f"Error calculating duration: {e}")
            
            row = {
                'User': username,
                'Session': session_label,
                'Start Time': start_time,
                'End Time': end_time,
                'Duration': duration,
                'Warning Count': sess.get('warning_count', 0)  # Use get with default to handle any edge cases
            }
            averages = sess.get('emotion_averages', {emotion: 0.0 for emotion in emotion_labels})
            for emotion in emotion_labels:
                row[f'{emotion} Average'] = averages.get(emotion, 0.0)
            data.append(row)
        
        # Create Excel file
        df = pd.DataFrame(data)
        
        # Reorder columns to make Warning Count more prominent (right after session info)
        columns = ['User', 'Session', 'Start Time', 'End Time', 'Duration', 'Warning Count']
        for emotion in emotion_labels:
            columns.append(f'{emotion} Average')
        
        # Ensure all columns exist in the DataFrame
        for col in columns:
            if col not in df.columns:
                df[col] = 0  # Add missing columns with default value
        
        # Reorder columns
        df = df[columns]
        
        # Sort by start time
        df = df.sort_values(by='Start Time', na_position='last')
        
        # Create a DataFrame specifically for emotion averages
        emotion_avg_data = []
        
        # Sort sessions by start_time to ensure consistent session numbering
        sorted_emotions = sorted(emotions, key=lambda x: x.get('start_time', '0'), reverse=False)
        
        # Create a dictionary to store session counts for each day
        session_counts = {}
        
        for sess in sorted_emotions:
            # Always recalculate emotion_averages to ensure it's up-to-date
            if 'emotions' in sess and sess['emotions']:
                # Recalculate emotion averages
                sess['emotion_averages'] = calculate_emotion_averages(sess['emotions'])
                logger.info(f"Calculated emotion averages for session {sess['session_id']}: {sess['emotion_averages']}")
            elif 'emotion_averages' not in sess:
                # Initialize with zeros if no emotions data
                sess['emotion_averages'] = {emotion: 0.0 for emotion in emotion_labels}
                logger.info(f"Initialized empty emotion averages for session {sess['session_id']}")
            
            # Save the updated emotion_averages back to the file
            with open(os.path.join(EMOTIONS_DIR, f"{username}_emotions.json"), 'w') as f:
                json.dump(sorted_emotions, f, indent=4)
            
            # Get the date from the start_time
            start_time = sess.get('start_time', 'N/A')
            if start_time != 'N/A':
                try:
                    # Extract the date part (YYYY-MM-DD)
                    session_date = start_time.split('T')[0]
                    
                    # Increment the session count for this date
                    if session_date not in session_counts:
                        session_counts[session_date] = 1
                    else:
                        session_counts[session_date] += 1
                    
                    # Format the session label as "Username - YYYY-MM-DD - Session #"
                    session_label = f"{username} - {session_date} - Session {session_counts[session_date]}"
                except Exception as e:
                    logger.error(f"Error parsing start_time {start_time}: {e}")
                    session_label = f"{username} - Session (time unknown)"
            else:
                session_label = f"{username} - Session (time unknown)"
            
            # Calculate session duration if both start and end times are available
            start_time = sess.get('start_time', 'N/A')
            end_time = sess.get('end_time', 'N/A')
            duration = 'N/A'
            
            if start_time != 'N/A' and end_time != 'N/A':
                try:
                    start_dt = datetime.fromisoformat(start_time)
                    end_dt = datetime.fromisoformat(end_time)
                    duration_seconds = (end_dt - start_dt).total_seconds()
                    
                    # Format duration as minutes and seconds
                    minutes = int(duration_seconds // 60)
                    seconds = int(duration_seconds % 60)
                    duration = f"{minutes}m {seconds}s"
                except Exception as e:
                    logger.error(f"Error calculating duration: {e}")
            
            row = {
                'User': username,
                'Session': session_label,
                'Start Time': start_time,
                'End Time': end_time,
                'Duration': duration,
                'Warning Count': sess.get('warning_count', 0)
            }
            
            # Get the emotion averages
            averages = sess.get('emotion_averages', {emotion: 0.0 for emotion in emotion_labels})
            logger.info(f"Emotion averages for {session_label}: {averages}")
            
            # Add each emotion average to the row
            for emotion in emotion_labels:
                row[emotion] = averages.get(emotion, 0.0)
            
            emotion_avg_data.append(row)
        
        emotion_df = pd.DataFrame(emotion_avg_data)
        emotion_df = emotion_df.sort_values(by='Start Time', na_position='last')
        
        # Create Excel file
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write the main data sheet
            df.to_excel(writer, index=False, sheet_name='Emotion Data', float_format='%.2f')
            
            # Write the emotion averages sheet
            emotion_df.to_excel(writer, index=False, sheet_name='Emotion Averages', float_format='%.2f')
            
            # Log the emotion averages data for debugging
            logger.info(f"Emotion averages data for {username}: {emotion_df.to_dict(orient='records')}")
            
            # Get the workbook and the worksheets
            workbook = writer.book
            worksheet = writer.sheets['Emotion Data']
            emotion_worksheet = writer.sheets['Emotion Averages']
            
            # Format the Warning Count column to stand out
            warning_count_idx = None
            for idx, col in enumerate(df.columns):
                if col == 'Warning Count':
                    warning_count_idx = idx
                    # Make column header bold with yellow background
                    cell = worksheet.cell(row=1, column=idx+1)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                    
                    # Set column width
                    worksheet.column_dimensions[get_column_letter(idx+1)].width = 15
                    
                    # Format all cells in the Warning Count column
                    for row_idx in range(2, len(df) + 2):  # +2 because Excel is 1-indexed and we have a header row
                        cell = worksheet.cell(row=row_idx, column=idx+1)
                        # If warning count is greater than 0, highlight it
                        if cell.value and int(cell.value) > 0:
                            cell.font = Font(bold=True, color='FF0000')  # Red text
                        cell.alignment = Alignment(horizontal='center')
            
            # Format the Emotion Averages sheet
            # Set column widths
            emotion_worksheet.column_dimensions['A'].width = 20  # User
            emotion_worksheet.column_dimensions['B'].width = 30  # Session
            emotion_worksheet.column_dimensions['C'].width = 20  # Start Time
            emotion_worksheet.column_dimensions['D'].width = 20  # End Time
            emotion_worksheet.column_dimensions['E'].width = 15  # Duration
            emotion_worksheet.column_dimensions['F'].width = 15  # Warning Count
            
            # Format headers
            for idx, col in enumerate(emotion_df.columns):
                cell = emotion_worksheet.cell(row=1, column=idx+1)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                
                # Color-code emotion headers
                if col in emotion_labels:
                    # Set column width
                    emotion_worksheet.column_dimensions[get_column_letter(idx+1)].width = 15
                    
                    # Set color based on emotion
                    color = 'DDDDDD'  # Default gray
                    if col == 'happy':
                        color = 'FFFF00'  # Yellow
                    elif col == 'sad':
                        color = 'ADD8E6'  # Light blue
                    elif col == 'angry':
                        color = 'FF9999'  # Light red
                    elif col == 'fear':
                        color = 'D8BFD8'  # Thistle
                    elif col == 'disgust':
                        color = '9ACD32'  # Yellow green
                    elif col == 'surprise':
                        color = 'FFA500'  # Orange
                    elif col == 'neutral':
                        color = 'E6E6E6'  # Light gray
                    
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    
                    # Format all cells in this emotion column
                    for row_idx in range(2, len(emotion_df) + 2):
                        cell = emotion_worksheet.cell(row=row_idx, column=idx+1)
                        value = cell.value
                        if value is not None:
                            # Format as percentage
                            cell.number_format = '0.00'
                            
                            # Highlight high values
                            if float(value) > 70:
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        
                        cell.alignment = Alignment(horizontal='center')
            
            # Create a summary sheet
            # Create a new sheet for the summary
            workbook.create_sheet('Warning Summary')
            summary_sheet = workbook['Warning Summary']
            
            # Add title
            summary_sheet['A1'] = f'Warning Summary for {username}'
            summary_sheet['A1'].font = Font(bold=True, size=14)
            summary_sheet.merge_cells('A1:F1')
            summary_sheet['A1'].alignment = Alignment(horizontal='center')
            
            # Add headers
            summary_sheet['A3'] = 'User'
            summary_sheet['B3'] = 'Session'
            summary_sheet['C3'] = 'Start Time'
            summary_sheet['D3'] = 'End Time'
            summary_sheet['E3'] = 'Duration'
            summary_sheet['F3'] = 'Warning Count'
            
            for col in range(1, 7):
                cell = summary_sheet.cell(row=3, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Add data - only sessions with warnings
            row_idx = 4
            for i, row in df.iterrows():
                warning_count = row['Warning Count']
                if warning_count and int(warning_count) > 0:
                    summary_sheet.cell(row=row_idx, column=1).value = row['User']
                    summary_sheet.cell(row=row_idx, column=2).value = row['Session']
                    summary_sheet.cell(row=row_idx, column=3).value = row['Start Time']
                    summary_sheet.cell(row=row_idx, column=4).value = row['End Time']
                    summary_sheet.cell(row=row_idx, column=5).value = row['Duration']
                    summary_sheet.cell(row=row_idx, column=6).value = warning_count
                    
                    # Format warning count
                    cell = summary_sheet.cell(row=row_idx, column=6)
                    cell.font = Font(bold=True, color='FF0000')
                    cell.alignment = Alignment(horizontal='center')
                    
                    row_idx += 1
            
            # Set column widths
            summary_sheet.column_dimensions['A'].width = 20  # User
            summary_sheet.column_dimensions['B'].width = 30  # Session
            summary_sheet.column_dimensions['C'].width = 20  # Start Time
            summary_sheet.column_dimensions['D'].width = 20  # End Time
            summary_sheet.column_dimensions['E'].width = 15  # Duration
            summary_sheet.column_dimensions['F'].width = 15  # Warning Count
            
            # Add total warnings
            total_warnings = df['Warning Count'].astype(int).sum()
            
            # Add a note if no warnings
            if row_idx == 4:  # No data was added
                summary_sheet.cell(row=4, column=1).value = "No sessions with warnings found"
                summary_sheet.merge_cells('A4:F4')
                summary_sheet.cell(row=4, column=1).alignment = Alignment(horizontal='center')
            
            # Add total row
            summary_sheet.cell(row=row_idx + 1, column=5).value = 'Total Warnings:'
            summary_sheet.cell(row=row_idx + 1, column=5).font = Font(bold=True)
            summary_sheet.cell(row=row_idx + 1, column=5).alignment = Alignment(horizontal='right')
            
            summary_sheet.cell(row=row_idx + 1, column=6).value = total_warnings
            summary_sheet.cell(row=row_idx + 1, column=6).font = Font(bold=True, color='FF0000')
            summary_sheet.cell(row=row_idx + 1, column=6).alignment = Alignment(horizontal='center')
            
            # Create an Emotion Summary sheet
            workbook.create_sheet('Emotion Summary')
            emotion_summary_sheet = workbook['Emotion Summary']
            
            # Add title
            emotion_summary_sheet['A1'] = f'Emotion Summary for {username}'
            emotion_summary_sheet['A1'].font = Font(bold=True, size=14)
            emotion_summary_sheet.merge_cells('A1:F1')
            emotion_summary_sheet['A1'].alignment = Alignment(horizontal='center')
            
            # Add headers
            emotion_summary_sheet['A3'] = 'User'
            emotion_summary_sheet['B3'] = 'Emotion'
            emotion_summary_sheet['C3'] = 'Average (%)'
            emotion_summary_sheet['D3'] = 'Max (%)'
            emotion_summary_sheet['E3'] = 'Min (%)'
            emotion_summary_sheet['F3'] = 'Sessions with High Values (>70%)'
            
            for col in range(1, 7):
                cell = emotion_summary_sheet.cell(row=3, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Calculate summary statistics for each emotion
            row_idx = 4
            for emotion in emotion_labels:
                # Get all values for this emotion
                values = []
                for sess in emotions:
                    # Ensure emotion_averages exists
                    if 'emotion_averages' not in sess and 'emotions' in sess and sess['emotions']:
                        sess['emotion_averages'] = calculate_emotion_averages(sess['emotions'])
                    
                    # Get the average for this emotion
                    avg = sess.get('emotion_averages', {}).get(emotion, 0.0)
                    logger.info(f"Session {sess['session_id']} - {emotion} average: {avg}")
                    
                    if avg > 0:  # Only include non-zero values
                        values.append(avg)
                
                logger.info(f"Values for {emotion}: {values}")
                
                if values:
                    avg_value = sum(values) / len(values)
                    max_value = max(values)
                    min_value = min(values)
                    high_count = sum(1 for v in values if v > 70)
                    logger.info(f"Summary for {emotion}: avg={avg_value}, max={max_value}, min={min_value}, high_count={high_count}")
                else:
                    avg_value = max_value = min_value = 0
                    high_count = 0
                    logger.info(f"No values for {emotion}")
                
                # Set color based on emotion
                color = 'DDDDDD'  # Default gray
                if emotion == 'happy':
                    color = 'FFFF00'  # Yellow
                elif emotion == 'sad':
                    color = 'ADD8E6'  # Light blue
                elif emotion == 'angry':
                    color = 'FF9999'  # Light red
                elif emotion == 'fear':
                    color = 'D8BFD8'  # Thistle
                elif emotion == 'disgust':
                    color = '9ACD32'  # Yellow green
                elif emotion == 'surprise':
                    color = 'FFA500'  # Orange
                elif emotion == 'neutral':
                    color = 'E6E6E6'  # Light gray
                
                # Add data
                emotion_summary_sheet.cell(row=row_idx, column=1).value = username
                emotion_summary_sheet.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center')
                
                emotion_summary_sheet.cell(row=row_idx, column=2).value = emotion.capitalize()
                emotion_summary_sheet.cell(row=row_idx, column=2).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                emotion_summary_sheet.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center')
                
                emotion_summary_sheet.cell(row=row_idx, column=3).value = avg_value
                emotion_summary_sheet.cell(row=row_idx, column=3).number_format = '0.00'
                emotion_summary_sheet.cell(row=row_idx, column=3).alignment = Alignment(horizontal='center')
                
                emotion_summary_sheet.cell(row=row_idx, column=4).value = max_value
                emotion_summary_sheet.cell(row=row_idx, column=4).number_format = '0.00'
                emotion_summary_sheet.cell(row=row_idx, column=4).alignment = Alignment(horizontal='center')
                
                emotion_summary_sheet.cell(row=row_idx, column=5).value = min_value
                emotion_summary_sheet.cell(row=row_idx, column=5).number_format = '0.00'
                emotion_summary_sheet.cell(row=row_idx, column=5).alignment = Alignment(horizontal='center')
                
                emotion_summary_sheet.cell(row=row_idx, column=6).value = high_count
                emotion_summary_sheet.cell(row=row_idx, column=6).alignment = Alignment(horizontal='center')
                
                row_idx += 1
            
            # Set column widths
            emotion_summary_sheet.column_dimensions['A'].width = 20  # User
            emotion_summary_sheet.column_dimensions['B'].width = 15  # Emotion
            emotion_summary_sheet.column_dimensions['C'].width = 15  # Average
            emotion_summary_sheet.column_dimensions['D'].width = 15  # Max
            emotion_summary_sheet.column_dimensions['E'].width = 15  # Min
            emotion_summary_sheet.column_dimensions['F'].width = 30  # Sessions with High Values
            
            # Set the Emotion Averages sheet as the active sheet when opening the file
            workbook.active = 1  # 0-based index, so 1 is the second sheet (Emotion Averages)
            
            # Add a title to each sheet
            for sheet_name in ['Emotion Data', 'Emotion Averages']:
                sheet = workbook[sheet_name]
                sheet.insert_rows(0)
                sheet.merge_cells('A1:Z1')  # Merge cells for the title
                title_cell = sheet.cell(row=1, column=1)
                title_cell.value = f"{sheet_name} for {username}"
                title_cell.font = Font(bold=True, size=14)
                title_cell.alignment = Alignment(horizontal='center')
            
            # Log the sheets in the workbook
            logger.info(f"Sheets in workbook: {workbook.sheetnames}")
        
        output.seek(0)
        logger.info(f"Generated Excel file for {username} with {len(data)} sessions, including warning counts")
        return send_file(
            output,
            download_name=f"{username}_emotions.xlsx",
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Error generating Excel for {username}: {e}")
        return jsonify({'success': False, 'message': 'Error generating Excel file'}), 500

@app.route('/video_feed')
def video_feed():
    if 'user_id' not in session or session.get('is_admin', False):
        logger.warning("Unauthorized access to video_feed")
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    logger.debug("Starting video feed")
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/get_emotion')
def get_emotion():
    if 'user_id' not in session or session.get('is_admin', False):
        logger.warning("Unauthorized access to get_emotion")
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    with lock:
        logger.debug(f"Returning emotion: {latest_emotion}")
        return jsonify({'success': True, 'data': latest_emotion})

@app.route('/get_session_averages')
def get_session_averages():
    if 'user_id' not in session or session.get('is_admin', False):
        logger.warning("Unauthorized access to get_session_averages")
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    username = session.get('username')
    session_id = session.get('session_id')
    
    if not username or not session_id:
        logger.warning("Missing username or session_id for get_session_averages")
        return jsonify({'success': False, 'message': 'Session data not found'}), 400
    
    try:
        # Load emotions data directly from file to ensure we have the latest data
        emotions_file = os.path.join(EMOTIONS_DIR, f"{username}_emotions.json")
        
        if os.path.exists(emotions_file):
            try:
                with open(emotions_file, 'r') as f:
                    emotions = json.load(f)
                logger.debug(f"Loaded emotions directly from file for {username}")
            except Exception as e:
                logger.error(f"Error loading emotions file directly: {e}")
                # Fall back to load_emotions function
                emotions = load_emotions(username)
        else:
            emotions = load_emotions(username)
        
        current_session = next((s for s in emotions if s.get('session_id') == session_id), None)
        
        if current_session and 'emotions' in current_session and current_session['emotions']:
            # Get the most recent emotions (last 10 or fewer)
            recent_emotions = current_session['emotions']
            if len(recent_emotions) > 10:
                recent_emotions = recent_emotions[-10:]  # Get the 10 most recent emotions
            
            # Calculate averages on the fly from recent emotions for more accurate real-time display
            averages = calculate_emotion_averages(recent_emotions)
            
            # Log the calculated averages for debugging
            logger.info(f"Calculated session averages for {username} from {len(recent_emotions)} recent emotions: {averages}")
            
            # Ensure all values are valid numbers
            for emotion in emotion_labels:
                if emotion not in averages or not isinstance(averages[emotion], (int, float)) or math.isnan(averages[emotion]):
                    averages[emotion] = 0.0
            
            return jsonify({
                'success': True, 
                'data': averages,
                'count': len(recent_emotions),
                'timestamp': datetime.now().isoformat()
            })
        else:
            # Return zeros if no emotions recorded yet
            empty_averages = {emotion: 0.0 for emotion in emotion_labels}
            logger.debug(f"No emotions found for session {session_id}, returning zeros")
            return jsonify({
                'success': True, 
                'data': empty_averages,
                'count': 0,
                'timestamp': datetime.now().isoformat()
            })
    except Exception as e:
        logger.error(f"Error getting session averages: {e}")
        return jsonify({'success': False, 'message': f'Error retrieving emotion averages: {str(e)}'}), 500

@app.route('/get_session_info')
def get_session_info():
    if 'user_id' not in session or session.get('is_admin', False):
        logger.warning("Unauthorized access to get_session_info")
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    username = session.get('username')
    session_id = session.get('session_id')
    
    if not username or not session_id:
        logger.warning("Missing username or session_id for get_session_info")
        return jsonify({'success': False, 'message': 'Session data not found'}), 400
    
    try:
        # Load emotions data directly from file to ensure we have the latest data
        emotions_file = os.path.join(EMOTIONS_DIR, f"{username}_emotions.json")
        
        if os.path.exists(emotions_file):
            try:
                with open(emotions_file, 'r') as f:
                    emotions = json.load(f)
                logger.debug(f"Loaded emotions directly from file for {username}")
            except Exception as e:
                logger.error(f"Error loading emotions file directly: {e}")
                # Fall back to load_emotions function
                emotions = load_emotions(username)
        else:
            emotions = load_emotions(username)
        
        current_session = next((s for s in emotions if s.get('session_id') == session_id), None)
        
        if current_session:
            # Ensure all warning counts exist
            updated = False
            
            # Check for sadness warning count
            if 'warning_count' not in current_session:
                current_session['warning_count'] = 0
                updated = True
            
            # Check for fear warning count
            if 'fear_warning_count' not in current_session:
                current_session['fear_warning_count'] = 0
                updated = True
                
            # Check for disgust warning count
            if 'disgust_warning_count' not in current_session:
                current_session['disgust_warning_count'] = 0
                updated = True
                
            # Check for anger warning count
            if 'angry_warning_count' not in current_session:
                current_session['angry_warning_count'] = 0
                updated = True
            
            # Save the updated session if any warning counts were added
            if updated:
                try:
                    with open(emotions_file, 'w') as f:
                        json.dump(emotions, f, indent=4)
                    logger.info(f"Added missing warning counts to session {session_id}")
                except Exception as e:
                    logger.error(f"Error saving updated session with warning counts: {e}")
            
            # Get all warning counts
            warning_count = current_session.get('warning_count', 0)
            fear_warning_count = current_session.get('fear_warning_count', 0)
            disgust_warning_count = current_session.get('disgust_warning_count', 0)
            angry_warning_count = current_session.get('angry_warning_count', 0)
            
            # Log the warning counts for debugging
            logger.info(f"Retrieved warning counts for {username}: sad={warning_count}, fear={fear_warning_count}, disgust={disgust_warning_count}, angry={angry_warning_count}")
            
            session_info = {
                'warning_count': warning_count,
                'fear_warning_count': fear_warning_count,
                'disgust_warning_count': disgust_warning_count,
                'angry_warning_count': angry_warning_count,
                'start_time': current_session.get('start_time'),
                'emotion_count': len(current_session.get('emotions', [])),
                'session_id': session_id,
                'last_updated': current_session.get('last_updated', 'Unknown')
            }
            
            logger.info(f"Returning session info for {username}: {session_info}")
            return jsonify({'success': True, 'data': session_info})
        else:
            logger.warning(f"Session {session_id} not found for {username}")
            return jsonify({'success': False, 'message': 'Session not found'}), 404
    except Exception as e:
        logger.error(f"Error getting session info: {e}")
        return jsonify({'success': False, 'message': f'Error retrieving session info: {str(e)}'}), 500

@app.route('/save_emotion', methods=['POST'])
def save_emotion():
    if 'user_id' not in session or session.get('is_admin', False):
        logger.warning("Unauthorized access to save_emotion")
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    data = request.get_json()
    emotion = data.get('emotion')
    confidence = data.get('confidence')
    username = session.get('username')
    session_id = session.get('session_id')
    logger.debug(f"Received save_emotion request: emotion={emotion}, confidence={confidence}, username={username}, session_id={session_id}")
    if not (emotion and confidence is not None and username and session_id):
        logger.error(f"Invalid emotion data: {data}, username: {username}, session_id: {session_id}")
        return jsonify({'success': False, 'message': 'Invalid data'}), 400
    if emotion == 'none':
        logger.debug(f"Skipping save for emotion='none'")
        return jsonify({'success': True, 'message': 'Skipped none emotion'})
    emotions = load_emotions(username)
    current_session = next((s for s in emotions if s['session_id'] == session_id), None)
    if not current_session:
        current_session = {
            'session_id': session_id,
            'start_time': session.get('start_time', datetime.now().isoformat()),
            'end_time': None,
            'emotions': [],
            'emotion_averages': {emotion: 0.0 for emotion in emotion_labels},
            'warning_count': 0  # Initialize warning count
        }
        logger.debug(f"Created new session for {username}: {session_id}")
    emotion_data = {
        'timestamp': datetime.now().isoformat(),
        'emotion': emotion,
        'confidence': float(confidence)
    }
    current_session['emotions'].append(emotion_data)
    try:
        save_emotions(username, current_session)
        logger.debug(f"Saved emotion to session {session_id} for {username}: {emotion_data}")
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Failed to save emotion for {username}: {e}")
        return jsonify({'success': False, 'message': 'Failed to save emotion'}), 500

@app.route('/save_warning', methods=['POST'])
def save_warning():
    if 'user_id' not in session or session.get('is_admin', False):
        logger.warning("Unauthorized access to save_warning")
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        data = request.get_json()
        logger.info(f"Received save_warning request with data: {data}")
        
        warning_count = data.get('warningCount')
        warning_type = data.get('warningType', 'sad')  # Default to 'sad' for backward compatibility
        username = session.get('username')
        session_id = session.get('session_id')
        
        logger.info(f"Processing save_warning: warning_type={warning_type}, warning_count={warning_count}, username={username}, session_id={session_id}")
        
        if warning_count is None or not username or not session_id:
            logger.error(f"Invalid warning data: {data}, username: {username}, session_id: {session_id}")
            return jsonify({'success': False, 'message': 'Invalid data'}), 400
        
        # Convert warning_count to integer if it's not already
        try:
            warning_count = int(warning_count)
        except (ValueError, TypeError):
            logger.error(f"Invalid warning count format: {warning_count}")
            return jsonify({'success': False, 'message': 'Invalid warning count format'}), 400
        
        # Load the user's emotion data
        emotions_file = os.path.join(EMOTIONS_DIR, f"{username}_emotions.json")
        if not os.path.exists(emotions_file):
            logger.error(f"Emotions file not found for {username}")
            return jsonify({'success': False, 'message': 'User data not found'}), 404
            
        # Load emotions directly from file to ensure we have the latest data
        try:
            with open(emotions_file, 'r') as f:
                emotions = json.load(f)
        except Exception as e:
            logger.error(f"Error loading emotions file for {username}: {e}")
            return jsonify({'success': False, 'message': 'Error loading user data'}), 500
            
        # Find the current session
        current_session = next((s for s in emotions if s.get('session_id') == session_id), None)
        
        if not current_session:
            logger.error(f"Session {session_id} not found for {username}")
            return jsonify({'success': False, 'message': 'Session not found'}), 404
        
        # Update the appropriate warning count based on warning_type
        if warning_type == 'sad':
            current_session['warning_count'] = warning_count
            logger.info(f"Setting sadness warning count to {warning_count} for session {session_id}")
        elif warning_type == 'fear':
            current_session['fear_warning_count'] = warning_count
            logger.info(f"Setting fear warning count to {warning_count} for session {session_id}")
        elif warning_type == 'disgust':
            current_session['disgust_warning_count'] = warning_count
            logger.info(f"Setting disgust warning count to {warning_count} for session {session_id}")
        elif warning_type == 'angry':
            current_session['angry_warning_count'] = warning_count
            logger.info(f"Setting anger warning count to {warning_count} for session {session_id}")
        else:
            logger.warning(f"Unknown warning type: {warning_type}, defaulting to sadness warning")
            current_session['warning_count'] = warning_count
        
        # Save the updated emotions data directly to file
        try:
            with open(emotions_file, 'w') as f:
                json.dump(emotions, f, indent=4)
            logger.info(f"Saved updated warning counts to {emotions_file}")
        except Exception as e:
            logger.error(f"Error saving updated warning counts to file: {e}")
            return jsonify({'success': False, 'message': 'Error saving warning counts'}), 500
        
        # Double-check that the file was updated correctly
        try:
            with open(emotions_file, 'r') as f:
                saved_data = json.load(f)
                saved_session = next((s for s in saved_data if s.get('session_id') == session_id), None)
                
                if not saved_session:
                    logger.error(f"Session {session_id} not found after save")
                    return jsonify({'success': False, 'message': 'Session not found after save'}), 500
                
                # Verify the appropriate warning count was saved
                if warning_type == 'sad':
                    saved_warning_count = saved_session.get('warning_count', 0)
                    if saved_warning_count != warning_count:
                        saved_session['warning_count'] = warning_count
                elif warning_type == 'fear':
                    saved_warning_count = saved_session.get('fear_warning_count', 0)
                    if saved_warning_count != warning_count:
                        saved_session['fear_warning_count'] = warning_count
                elif warning_type == 'disgust':
                    saved_warning_count = saved_session.get('disgust_warning_count', 0)
                    if saved_warning_count != warning_count:
                        saved_session['disgust_warning_count'] = warning_count
                elif warning_type == 'angry':
                    saved_warning_count = saved_session.get('angry_warning_count', 0)
                    if saved_warning_count != warning_count:
                        saved_session['angry_warning_count'] = warning_count
                
                # If any warning count was incorrect, force an update
                with open(emotions_file, 'w') as f2:
                    json.dump(saved_data, f2, indent=4)
                logger.info(f"Verified warning counts for session {session_id}")
                
        except Exception as e:
            logger.error(f"Error verifying saved warning counts: {e}")
            # Continue anyway since we already saved once
        
        # Prepare response with all warning counts
        response_data = {
            'success': True,
            'message': f'{warning_type.capitalize()} warning count updated to {warning_count}',
            'warning_count': current_session.get('warning_count', 0),
            'fear_warning_count': current_session.get('fear_warning_count', 0),
            'disgust_warning_count': current_session.get('disgust_warning_count', 0),
            'angry_warning_count': current_session.get('angry_warning_count', 0)
        }
        
        logger.info(f"Successfully updated {warning_type} warning count for {username} to {warning_count}")
        return jsonify(response_data)
    
    except Exception as e:
        logger.error(f"Failed to save warning count: {str(e)}")
        return jsonify({'success': False, 'message': f'Failed to save warning count: {str(e)}'}), 500

@app.route('/stop', methods=['GET'])
def stop():
    global cap, stop_camera
    logger.debug("Stop route called")
    stop_camera = True
    if cap is not None:
        cap.release()
        logger.debug("Camera released")
    return redirect(url_for('home'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.get_json()
        if not data or 'email' not in data or 'password' not in data:
            logger.error("Invalid login request data")
            return jsonify({'success': False, 'message': 'Missing email or password'}), 400
        email = data['email'].strip()
        password = data['password']
        if not email or not password:
            logger.warning("Empty login fields")
            return jsonify({'success': False, 'message': 'Email and password are required'}), 400
        logger.debug(f"Login attempt with email: {email}")
        if email.lower() == 'admin':
            logger.debug("Processing admin login")
            if pbkdf2_sha256.verify(password, ADMIN_PASSWORD_HASH):
                session.clear()
                session['is_admin'] = True
                session.permanent = True
                logger.info("Admin login successful")
                return jsonify({'success': True, 'redirect': url_for('admin_dashboard', _external=True)})
            else:
                logger.warning("Admin login failed: incorrect password")
                return jsonify({'success': False, 'message': 'Invalid admin credentials'})
        users = load_users()
        user_data = users.get(email)
        if user_data:
            logger.debug(f"User found for email: {email}")
            if pbkdf2_sha256.verify(password, user_data['password']):
                # Clear previous session and create a new one
                session.clear()
                session['user_id'] = user_data['user_id']
                session['username'] = user_data['username']
                session['email'] = email
                session['session_id'] = str(uuid4())
                session['start_time'] = datetime.now().isoformat()
                session.permanent = True
                
                # Create a new session record in the user's emotion file
                new_session = {
                    'session_id': session['session_id'],
                    'start_time': session['start_time'],
                    'end_time': None,
                    'emotions': [],
                    'emotion_averages': {emotion: 0.0 for emotion in emotion_labels},
                    'warning_count': 0,
                    'user_agent': request.headers.get('User-Agent', 'Unknown'),
                    'ip_address': request.remote_addr,
                    'login_type': 'regular'
                }
                
                try:
                    # Save the new session to the user's emotion file
                    save_emotions(user_data['username'], new_session)
                    logger.info(f"Created new session for {user_data['username']}, session_id: {session['session_id']}")
                except Exception as e:
                    logger.error(f"Failed to create new session for {user_data['username']}: {e}")
                
                logger.info(f"User {user_data['username']} logged in, session_id: {session['session_id']}")
                return jsonify({'success': True, 'redirect': url_for('dashboard', _external=True)})
            else:
                logger.warning(f"User login failed for email: {email} - incorrect password")
                return jsonify({'success': False, 'message': 'Invalid email or password'})
        else:
            logger.warning(f"User login failed for email: {email} - user not found")
            return jsonify({'success': False, 'message': 'Invalid email or password'})
    logger.debug("Rendering login page")
    return render_template('log.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        data = request.get_json()
        if not data or 'username' not in data or 'email' not in data or 'password' not in data:
            logger.error("Invalid signup request data")
            return jsonify({'success': False, 'message': 'Missing username, email, or password'}), 400
        username = data['username'].strip()
        email = data['email'].strip()
        password = data['password']
        if not username or not email or not password:
            logger.warning("Empty signup fields")
            return jsonify({'success': False, 'message': 'All fields are required'}), 400
        users = load_users()
        if email in users:
            logger.warning(f"Email already registered: {email}")
            return jsonify({'success': False, 'message': 'Email already registered'})
        user_id = str(uuid4())
        hashed_password = pbkdf2_sha256.hash(password)
        users[email] = {
            'user_id': user_id,
            'username': username,
            'password': hashed_password
        }
        save_users(users)
        
        # Clear previous session and create a new one
        session.clear()
        session['user_id'] = user_id
        session['username'] = username
        session['email'] = email
        session['session_id'] = str(uuid4())
        session['start_time'] = datetime.now().isoformat()
        session.permanent = True
        
        # Create a new session record in the user's emotion file
        new_session = {
            'session_id': session['session_id'],
            'start_time': session['start_time'],
            'end_time': None,
            'emotions': [],
            'emotion_averages': {emotion: 0.0 for emotion in emotion_labels},
            'warning_count': 0,
            'user_agent': request.headers.get('User-Agent', 'Unknown'),
            'ip_address': request.remote_addr,
            'login_type': 'signup',  # Mark this as a signup session
            'first_session': True    # Flag to indicate this is the user's first session
        }
        
        try:
            # Save the new session to the user's emotion file
            save_emotions(username, new_session)
            logger.info(f"Created first session for new user {username}, session_id: {session['session_id']}")
        except Exception as e:
            logger.error(f"Failed to create first session for new user {username}: {e}")
        
        logger.info(f"User {username} signed up, session_id: {session['session_id']}")
        return jsonify({'success': True, 'redirect': url_for('dashboard', _external=True)})
    logger.debug("Rendering signup page")
    return render_template('sign.html')

@app.route('/logout')
def logout():
    global cap, stop_camera
    logger.debug("Logout route called")
    stop_camera = True
    if cap is not None:
        cap.release()
        logger.debug("Camera released on logout")
    if 'user_id' in session and 'username' in session and 'session_id' in session:
        username = session['username']
        session_id = session['session_id']
        emotions = load_emotions(username)
        current_session = next((s for s in emotions if s['session_id'] == session_id), None)
        
        if current_session:
            # Record session end time
            end_time = datetime.now().isoformat()
            current_session['end_time'] = end_time
            
            # Calculate emotion averages if there are emotions recorded
            if current_session.get('emotions', []):
                current_session['emotion_averages'] = calculate_emotion_averages(current_session['emotions'])
            
            # Calculate session duration
            if 'start_time' in current_session:
                try:
                    start_dt = datetime.fromisoformat(current_session['start_time'])
                    end_dt = datetime.fromisoformat(end_time)
                    duration_seconds = (end_dt - start_dt).total_seconds()
                    current_session['duration_seconds'] = duration_seconds
                    
                    # Format duration as minutes and seconds
                    minutes = int(duration_seconds // 60)
                    seconds = int(duration_seconds % 60)
                    current_session['duration_formatted'] = f"{minutes}m {seconds}s"
                    
                    logger.info(f"Session duration for {username}: {current_session['duration_formatted']}")
                except Exception as e:
                    logger.error(f"Error calculating session duration on logout: {e}")
            
            # Add logout information
            current_session['logout_type'] = 'normal'
            current_session['logout_time'] = end_time
            
            try:
                # Save the updated session data
                save_emotions(username, current_session)
                logger.debug(f"Session {session_id} for {username} ended at {end_time}, "
                            f"Duration: {current_session.get('duration_formatted', 'N/A')}, "
                            f"Emotion Averages: {current_session['emotion_averages']}, "
                            f"Warning Count: {current_session.get('warning_count', 0)}")
                
                # Generate and save Excel file with session data
                try:
                    excel_file_path = os.path.join(EMOTIONS_DIR, f"{username}_sessions.xlsx")
                    create_user_excel_report(username, excel_file_path)
                    logger.info(f"Generated Excel report for {username} at {excel_file_path}")
                except Exception as e:
                    logger.error(f"Failed to generate Excel report on logout for {username}: {e}")
                    
            except Exception as e:
                logger.error(f"Failed to save session on logout for {username}: {e}")
    session.clear()
    logger.info("User or admin logged out")
    return redirect(url_for('home'))

@app.route('/debug/routes')
def debug_routes():
    routes = [str(rule) for rule in app.url_map.iter_rules()]
    logger.debug(f"Registered routes: {routes}")
    return jsonify({'routes': routes})

if __name__ == '__main__':
    if not os.path.exists('templates'):
        os.makedirs('templates')
    if not os.path.exists('static'):
        os.makedirs('static')
    logger.info("Starting Flask application")
    app.run(debug=True, host='0.0.0.0', port=5000, use_reloader=False)